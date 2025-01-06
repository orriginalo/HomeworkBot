import openpyxl
from openpyxl.styles import PatternFill, Alignment
from rich import print

def format_table():
    # Открываем существующий файл Excel
    file_path = "domashkabot info.xlsx"
    wb = openpyxl.load_workbook(file_path)

    # Функция для настройки листа
    def format_sheet(sheet, header_color, column_settings):
        # Устанавливаем цвет заголовка
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        
        # Форматируем заголовок
        for col in range(1, len(column_settings) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Устанавливаем ширину столбцов и выравнивание
        for col_index, down_settings in enumerate(column_settings.items(), start=1):
            settings = down_settings[1]
            # print(col_index ,settings[0], settings[1])  # Вывод настроек ширины и выравнивания
            
            # Устанавливаем ширину столбца
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = settings[0] / 7
            # print(f"Устанавливаем ширину для столбца {col_index}: {settings[0]}")  # Для отладки
            
            if settings[1]:  # Если нужно выравнивание по центру
                for row in range(1, sheet.max_row + 1):  # Применяем выравнивание ко всем строкам
                    sheet.cell(row=row, column=col_index).alignment = Alignment(horizontal='center', vertical='center')
        
        # print(f"Отформатирован лист: {sheet.title}\n")  # Для отладки

    # Лист "Домашние задания"
    if "Домашние задания" in wb.sheetnames:
        sheet1 = wb["Домашние задания"]
        format_sheet(sheet1, '92D050', {  # Зеленый цвет
            1: (38, True),
            2: (147, True),
            3: (171, True),
            4: (450, False),
            5: (105, True)
        })
    else:
        # print("Лист 'Домашние задания' не найден.")
        pass

    # Лист "Пользователи"
    if "Пользователи" in wb.sheetnames:
        sheet2 = wb["Пользователи"]
        format_sheet(sheet2, 'ADD8E6', {  # Светло-синий цвет
            1: (108, True),
            2: (130, True),
            3: (51, True),
            4: (120, True)
        })
    else:
        # print("Лист 'Пользователи' не найден.")
        pass

    # Лист "Расписание пар"
    if "Расписание пар" in wb.sheetnames:
        sheet3 = wb["Расписание пар"]
        format_sheet(sheet3, 'E6E6FA', {  # Светло-фиолетовый цвет
            1: (102, True),
            2: (178, False),
            3: (113, True)
        })
    else:
        # print("Лист 'Расписание пар' не найден.")
        pass

    # Сохраняем изменения в файле Excel
    wb.save(file_path)

# format_table()